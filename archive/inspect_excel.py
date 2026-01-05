import openpyxl

FILE = "Name List for Canteen e-Recipts.xlsx"

try:
    wb = openpyxl.load_workbook(FILE)
    ws = wb.active
    print(f"Sheet Name: {ws.title}")
    
    # Get Headers (First Row)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value))
    
    print(f"Headers: {headers}")
    
    # Print first few rows of data to check content
    print("First 3 rows of data:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
        print(f"Row {i+1}: {row}")

except Exception as e:
    print(f"Error reading file: {e}")
